"""
Excel export module — generates a professionally formatted, color-coded
Excel spreadsheet from Gemini analysis results.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

from models.schemas import GeminiAnalysisResult


# ─── Color palette ───
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

STRONG_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # green-100
POSSIBLE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # yellow-100
WEAK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")      # red-100

STRONG_FONT = Font(name="Calibri", size=10, bold=True, color="065F46")
POSSIBLE_FONT = Font(name="Calibri", size=10, bold=True, color="92400E")
WEAK_FONT = Font(name="Calibri", size=10, bold=True, color="991B1B")

SECTION_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="E5E7EB")

NORMAL_FONT = Font(name="Calibri", size=10, color="1F2937")
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="1F2937")
SCORE_FONT_HIGH = Font(name="Calibri", size=10, bold=True, color="065F46")
SCORE_FONT_MED = Font(name="Calibri", size=10, bold=True, color="92400E")
SCORE_FONT_LOW = Font(name="Calibri", size=10, bold=True, color="991B1B")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

LIGHT_GRAY_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

RED_FLAG_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
RED_FLAG_FONT = Font(name="Calibri", size=10, color="991B1B")


def _get_tier_style(tier: str):
    """Return fill and font for fit tier."""
    tier_upper = tier.strip().lower()
    if "strong" in tier_upper:
        return STRONG_FILL, STRONG_FONT
    elif "possible" in tier_upper:
        return POSSIBLE_FILL, POSSIBLE_FONT
    else:
        return WEAK_FILL, WEAK_FONT


def _get_score_font(score: float) -> Font:
    if score >= 70:
        return SCORE_FONT_HIGH
    elif score >= 45:
        return SCORE_FONT_MED
    else:
        return SCORE_FONT_LOW


def _get_score_fill(score: float) -> PatternFill:
    if score >= 70:
        return STRONG_FILL
    elif score >= 45:
        return POSSIBLE_FILL
    else:
        return WEAK_FILL


def generate_excel_report(
    analyses: List[Dict[str, Any]],
    job_title: str,
    job_description: str,
    output_dir: str
) -> str:
    """
    Generate a color-coded Excel report from a list of candidate analyses.
    
    Parameters:
        analyses: List of dicts with keys 'candidate_name', 'analysis' (GeminiAnalysisResult dict)
        job_title: The job title
        job_description: The job description
        output_dir: Directory to save the file
    
    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    
    # ════════════════════════════════════════════
    # Sheet 1: Summary / Leaderboard
    # ════════════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Candidate Summary"
    ws_summary.sheet_properties.tabColor = "4F46E5"
    
    # Title row
    ws_summary.merge_cells("A1:L1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"AI Hiring Report — {job_title}"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # Subtitle
    ws_summary.merge_cells("A2:L2")
    subtitle = ws_summary["A2"]
    subtitle.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    subtitle.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    subtitle.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        "Rank", "Candidate Name", "Email", "Phone", "Location",
        "Overall Score", "Fit Tier", "Skills Score", "Experience Score",
        "Semantic Score", "Verdict", "Recommended Action"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws_summary.row_dimensions[4].height = 30
    
    # Sort by overall score descending
    sorted_analyses = sorted(
        analyses,
        key=lambda x: x.get("analysis", {}).get("match", {}).get("overall_score_percent", 0),
        reverse=True
    )
    
    for row_idx, item in enumerate(sorted_analyses, 5):
        a = item.get("analysis", {})
        candidate = a.get("candidate", {})
        match = a.get("match", {})
        
        rank = row_idx - 4
        overall_score = match.get("overall_score_percent", 0)
        fit_tier = match.get("fit_tier", "Weak")
        tier_fill, tier_font = _get_tier_style(fit_tier)
        
        row_fill = LIGHT_GRAY_FILL if rank % 2 == 0 else WHITE_FILL
        
        values = [
            rank,
            candidate.get("full_name", item.get("candidate_name", "Unknown")),
            candidate.get("email", "Not specified"),
            candidate.get("phone", "Not specified"),
            candidate.get("location", "Not specified"),
            f"{overall_score}%",
            fit_tier,
            f"{match.get('skills_score_percent', 0)}%",
            f"{match.get('experience_score_percent', 0)}%",
            f"{match.get('semantic_relevance_score_percent', 0)}%",
            match.get("one_line_verdict", ""),
            a.get("recruiter_tools", {}).get("recommended_action", ""),
        ]
        
        for col_idx, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        
        # Color the score cell
        score_cell = ws_summary.cell(row=row_idx, column=6)
        score_cell.font = _get_score_font(overall_score)
        score_cell.fill = _get_score_fill(overall_score)
        
        # Color the tier cell
        tier_cell = ws_summary.cell(row=row_idx, column=7)
        tier_cell.fill = tier_fill
        tier_cell.font = tier_font
        
        # Color individual scores
        for col in [8, 9, 10]:
            c = ws_summary.cell(row=row_idx, column=col)
            try:
                sv = float(str(c.value).replace('%', ''))
                c.font = _get_score_font(sv)
            except:
                pass
        
        ws_summary.row_dimensions[row_idx].height = 35
    
    # Auto-fit column widths
    col_widths = [6, 25, 28, 18, 20, 14, 12, 14, 16, 14, 40, 22]
    for i, w in enumerate(col_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w
    
    # ════════════════════════════════════════════
    # Sheet 2: Detailed Analysis per candidate
    # ════════════════════════════════════════════
    ws_detail = wb.create_sheet("Detailed Analysis")
    ws_detail.sheet_properties.tabColor = "7C3AED"
    
    detail_row = 1
    
    for item in sorted_analyses:
        a = item.get("analysis", {})
        candidate = a.get("candidate", {})
        match = a.get("match", {})
        experience = a.get("experience", {})
        education = a.get("education", {})
        skills = a.get("skills", {})
        red_flags = a.get("red_flags", {})
        recruiter = a.get("recruiter_tools", {})
        
        name = candidate.get("full_name", item.get("candidate_name", "Unknown"))
        
        # ── Candidate header ──
        ws_detail.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=6)
        header_cell = ws_detail.cell(row=detail_row, column=1, value=f"📋 {name}")
        header_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(2, 7):
            ws_detail.cell(row=detail_row, column=c).fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        ws_detail.row_dimensions[detail_row].height = 32
        detail_row += 1
        
        # ── Contact Info ──
        detail_row = _write_section_header(ws_detail, detail_row, "Contact Information")
        contact_fields = [
            ("Email", candidate.get("email", "Not specified")),
            ("Phone", candidate.get("phone", "Not specified")),
            ("Location", candidate.get("location", "Not specified")),
            ("LinkedIn", candidate.get("linkedin_url", "Not specified")),
            ("Portfolio", candidate.get("portfolio_url", "Not specified")),
        ]
        for label, value in contact_fields:
            detail_row = _write_field(ws_detail, detail_row, label, value)
        
        # ── Match Scores ──
        detail_row = _write_section_header(ws_detail, detail_row, "Match Scores")
        overall = match.get("overall_score_percent", 0)
        detail_row = _write_score_field(ws_detail, detail_row, "Overall Score", overall)
        detail_row = _write_field(ws_detail, detail_row, "Fit Tier", match.get("fit_tier", "Weak"), tier_style=True)
        detail_row = _write_field(ws_detail, detail_row, "Verdict", match.get("one_line_verdict", ""))
        detail_row = _write_score_field(ws_detail, detail_row, "Skills Score", match.get("skills_score_percent", 0))
        detail_row = _write_score_field(ws_detail, detail_row, "Experience Score", match.get("experience_score_percent", 0))
        detail_row = _write_score_field(ws_detail, detail_row, "Semantic Relevance", match.get("semantic_relevance_score_percent", 0))
        
        # ── Experience ──
        detail_row = _write_section_header(ws_detail, detail_row, "Experience")
        exp_fields = [
            ("Total Years", str(experience.get("total_years", 0))),
            ("Current Role", experience.get("current_role", "Not specified")),
            ("Current Company", experience.get("current_company", "Not specified")),
            ("Career Progression", experience.get("career_progression", "Not specified")),
            ("Industry", experience.get("industry_background", "Not specified")),
            ("Notable Companies", ", ".join(experience.get("notable_companies", []))),
        ]
        for label, value in exp_fields:
            detail_row = _write_field(ws_detail, detail_row, label, value)
        
        # ── Education ──
        detail_row = _write_section_header(ws_detail, detail_row, "Education")
        edu_fields = [
            ("Highest Degree", education.get("highest_degree", "Not specified")),
            ("Field of Study", education.get("field_of_study", "Not specified")),
            ("Institution", education.get("institution", "Not specified")),
            ("Graduation Year", education.get("graduation_year", "Not specified")),
        ]
        for label, value in edu_fields:
            detail_row = _write_field(ws_detail, detail_row, label, value)
        
        # ── Skills ──
        detail_row = _write_section_header(ws_detail, detail_row, "Skills Analysis")
        detail_row = _write_field(ws_detail, detail_row, "Matched Skills", ", ".join(skills.get("matched_required_skills", [])))
        detail_row = _write_field(ws_detail, detail_row, "Missing Skills", ", ".join(skills.get("missing_required_skills", [])), highlight_red=True)
        detail_row = _write_field(ws_detail, detail_row, "Bonus Skills", ", ".join(skills.get("bonus_skills", [])))
        detail_row = _write_field(ws_detail, detail_row, "Tech Stack", skills.get("tech_stack_summary", "Not specified"))
        
        # ── Red Flags ──
        detail_row = _write_section_header(ws_detail, detail_row, "⚠️ Red Flags")
        rf_fields = [
            ("Employment Gaps", red_flags.get("employment_gaps", "None")),
            ("Job Hopping", red_flags.get("job_hopping", "None")),
            ("Role Mismatch", red_flags.get("role_mismatch", "None")),
            ("Other Concerns", red_flags.get("other_concerns", "")),
        ]
        for label, value in rf_fields:
            is_flagged = value and value.lower() not in ("none", "", "not specified")
            detail_row = _write_field(ws_detail, detail_row, label, value or "None", highlight_red=is_flagged)
        
        # ── Recruiter Tools ──
        detail_row = _write_section_header(ws_detail, detail_row, "Recruiter Insights")
        detail_row = _write_field(ws_detail, detail_row, "Strengths", recruiter.get("strengths_summary", ""))
        detail_row = _write_field(ws_detail, detail_row, "Weaknesses", recruiter.get("weaknesses_summary", ""))
        detail_row = _write_field(ws_detail, detail_row, "Action", recruiter.get("recommended_action", ""))
        
        questions = recruiter.get("suggested_interview_questions", [])
        for i, q in enumerate(questions, 1):
            detail_row = _write_field(ws_detail, detail_row, f"Interview Q{i}", q)
        
        # Spacer
        detail_row += 2
    
    # Set column widths for detail sheet
    ws_detail.column_dimensions["A"].width = 4
    ws_detail.column_dimensions["B"].width = 22
    ws_detail.column_dimensions["C"].width = 60
    ws_detail.column_dimensions["D"].width = 15
    ws_detail.column_dimensions["E"].width = 15
    ws_detail.column_dimensions["F"].width = 15
    
    # ════════════════════════════════════════════
    # Sheet 3: Skills Matrix
    # ════════════════════════════════════════════
    ws_skills = wb.create_sheet("Skills Matrix")
    ws_skills.sheet_properties.tabColor = "059669"
    
    # Collect all unique required skills across all candidates
    all_required_skills = set()
    for item in sorted_analyses:
        a = item.get("analysis", {})
        skills_data = a.get("skills", {})
        all_required_skills.update(skills_data.get("matched_required_skills", []))
        all_required_skills.update(skills_data.get("missing_required_skills", []))
    
    all_required_skills = sorted(all_required_skills)
    
    # Header row
    ws_skills.cell(row=1, column=1, value="Candidate").font = HEADER_FONT
    ws_skills.cell(row=1, column=1).fill = HEADER_FILL
    ws_skills.cell(row=1, column=1).border = THIN_BORDER
    
    for col_idx, skill in enumerate(all_required_skills, 2):
        cell = ws_skills.cell(row=1, column=col_idx, value=skill)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, text_rotation=45)
        cell.border = THIN_BORDER
    
    ws_skills.row_dimensions[1].height = 50
    
    for row_idx, item in enumerate(sorted_analyses, 2):
        a = item.get("analysis", {})
        name = a.get("candidate", {}).get("full_name", item.get("candidate_name", "Unknown"))
        matched = [s.lower() for s in a.get("skills", {}).get("matched_required_skills", [])]
        
        name_cell = ws_skills.cell(row=row_idx, column=1, value=name)
        name_cell.font = BOLD_FONT
        name_cell.border = THIN_BORDER
        
        for col_idx, skill in enumerate(all_required_skills, 2):
            cell = ws_skills.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            
            if skill.lower() in matched:
                cell.value = "✓"
                cell.font = Font(name="Calibri", size=12, bold=True, color="065F46")
                cell.fill = STRONG_FILL
            else:
                cell.value = "✗"
                cell.font = Font(name="Calibri", size=12, bold=True, color="991B1B")
                cell.fill = WEAK_FILL
    
    ws_skills.column_dimensions["A"].width = 25
    for i in range(2, len(all_required_skills) + 2):
        ws_skills.column_dimensions[get_column_letter(i)].width = 14
    
    # Save
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"hiring_report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath


# ─── Helper functions for detailed sheet ───

def _write_section_header(ws, row: int, title: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    ws.row_dimensions[row].height = 24
    return row + 1


def _write_field(ws, row: int, label: str, value: str, tier_style: bool = False, highlight_red: bool = False) -> int:
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = BOLD_FONT
    label_cell.border = THIN_BORDER
    label_cell.alignment = Alignment(vertical="center")
    
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    value_cell = ws.cell(row=row, column=3, value=value or "Not specified")
    value_cell.alignment = Alignment(wrap_text=True, vertical="center")
    value_cell.border = THIN_BORDER
    
    if tier_style and value:
        fill, font = _get_tier_style(value)
        value_cell.fill = fill
        value_cell.font = font
    elif highlight_red:
        value_cell.fill = RED_FLAG_FILL
        value_cell.font = RED_FLAG_FONT
    else:
        value_cell.font = NORMAL_FONT
    
    return row + 1


def _write_score_field(ws, row: int, label: str, score: float) -> int:
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = BOLD_FONT
    label_cell.border = THIN_BORDER
    
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    value_cell = ws.cell(row=row, column=3, value=f"{score}%")
    value_cell.font = _get_score_font(score)
    value_cell.fill = _get_score_fill(score)
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.border = THIN_BORDER
    
    return row + 1
